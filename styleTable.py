import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
import pandas as pd
import pyperclip
import io
import json
import re
from math import ceil

def load_config(config_file='table_export_config.json'):
    with open(config_file, 'r') as f:
        return json.load(f)

def rgb_to_hex(rgb):
    if rgb is None:
        return None
    if isinstance(rgb, str):
        return rgb.replace('00', '')  # Remove leading '00' if present
    if hasattr(rgb, 'rgb'):
        return rgb.rgb.replace('00', '')  # Remove leading '00' if present
    if isinstance(rgb, tuple) and len(rgb) == 3:
        return '{:02x}{:02x}{:02x}'.format(*rgb)
    return None

def evaluate_formula(formula, worksheet):
    formula = formula[1:]
    
    total = 0
    current_operator = '+'
    
    for token in formula.split():
        if token in ['+', '-']:
            current_operator = token
        elif re.match(r'[A-Z]+\d+', token):  # If the token is a cell reference
            cell_value = worksheet[token].value
            
            # Handle cases where the cell value is None or not a number
            if cell_value is None:
                cell_value = 0  # Treat None as 0 for evaluation
            elif not isinstance(cell_value, (int, float)):
                raise ValueError(f"Cell {token} has a non-numeric value: {cell_value}")

            # Apply the operator
            if current_operator == '+':
                total += cell_value
            elif current_operator == '-':
                total -= cell_value
    return total

# Function to calculate the height needed based on text length and column width
def calculate_row_height(cell_value, width, wrap_threshold, font_size):
    if not cell_value:
        return 15.0  # Default row height if the cell is empty
    lines = ceil(len(str(cell_value)) / wrap_threshold)  # Estimate number of lines
    return float(lines * (font_size + 2))  # Adjust row height based on font size
        
def excel_to_html_with_style(xlsx_file_path, config, row_skip, output_excel_file="output.xlsx"):
    # Load the existing workbook
    wb = openpyxl.load_workbook(xlsx_file_path)
    ws = wb.active  
    for row in range(1, ws.max_row + 1):  # Start from row 1 to include all rows
        cell = ws.cell(row=row, column=1)  # Access the cell in column 1
        cell.value = None  # Empty the cell value
        cell.border = Border()  # Set border to none
    
    # Read the data into a DataFrame (without loading formulas for now)
    data = ws.values
    columns = next(data)[0:]  # Get the first row for the column headers
    main_df = pd.DataFrame(data, columns=columns)

    # Apply styles based on configuration
    header_fill = PatternFill(
        start_color=config['header']['backgroundColor'],
        end_color=config['header']['backgroundColor'],
        fill_type='solid'
    )
    header_font = Font(
        color=config['header']['fontColor'],
        bold=config['header'].get('bold', False),
        size=config['header']['fontSize'],
        name=config['font']['family']
    )
    body_font = Font(
        color=config['body']['fontColor'],
        size=config['body']['fontSize'],
        name=config['font']['family']
    )
    alignment = Alignment(
        horizontal=config['alignment']['horizontal'],
        vertical=config['alignment']['vertical'],
        wrap_text=config['alignment'].get('wrapText', True)
    )
    header_alignment = Alignment(
        horizontal=config['headeralignment']['horizontal'],
        vertical=config['headeralignment']['vertical'],
        wrap_text=config['headeralignment'].get('wrapText', True)
    )

    # Handle borders based on configuration
    if config['cell'].get('borders', True):
        border = Border(
            left=Side(style=config['cell']['borderStyle'], color=config['cell']['borderColor']),
            right=Side(style=config['cell']['borderStyle'], color=config['cell']['borderColor']),
            top=Side(style=config['cell']['borderStyle'], color=config['cell']['borderColor']),
            bottom=Side(style=config['cell']['borderStyle'], color=config['cell']['borderColor'])
        )
    else:
        border = Border()

    # Set column widths based on max length of content in each column, considering merged cells
    widths=[]
    for c, column in enumerate(main_df.columns[1:], start=2):  # Skip the first column
        max_length = 0
        for r in range(len(main_df)):
            cell_value = main_df[column].iloc[r]
            if isinstance(cell_value, str):
                max_length = max(max_length, len(cell_value))
            elif isinstance(cell_value, (int, float)):
                max_length = max(max_length, len(str(cell_value)))
            elif isinstance(cell_value, (pd.Timestamp)):  # Check for date types
                max_length = max(max_length, len(cell_value.strftime('%Y-%m-%d')))  # Format date as YYYY-MM-DD

            # Check for currency or percentage formatting
            if isinstance(cell_value, str):
                if cell_value.startswith('$') or cell_value.endswith('%'):
                    max_length = max(max_length, len(cell_value))

        # Calculate the final value of the formula if it exists
        if isinstance(cell_value, str) and cell_value.startswith('='):
            # Evaluate the formula to get the final value
            try:
                final_value = eval(cell_value[1:])  # Remove '=' and evaluate
                max_length = max(max_length, len(str(final_value)))
            except Exception:
                pass  # If evaluation fails, ignore

        # Adjust width for merged cells by checking the leftmost cell in the merged range
        for merged_cell_range in ws.merged_cells.ranges:
            if merged_cell_range.min_col <= c <= merged_cell_range.max_col:  # Check if the merged range includes the current column
                max_length = max(max_length, len(str(ws.cell(row=merged_cell_range.min_row, column=merged_cell_range.min_col).value)))

        adjusted_width = min(max_length + 5, config['cell']['wrapThreshold'])  # Adding extra space for padding
        widths.append(adjusted_width)
        ws.column_dimensions[get_column_letter(c)].width = adjusted_width  # Set column width in Excel

    # Apply styles and adjust row heights for merged header cells
    for merged_cell_range in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged_cell_range.min_row, merged_cell_range.min_col, merged_cell_range.max_row, merged_cell_range.max_col

        # Apply styles to all the cells in the merged header range
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if col==1: continue
                cell = ws.cell(row=row, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

        # Calculate the required height for the entire row (only for the first row of the merged range)

    # For non-merged header cells (just in case)
    
    for row in range(1, row_skip + 1):
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if not any(cell.coordinate in merged_cell_range for merged_cell_range in ws.merged_cells.ranges):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

                # Adjust height based on text wrapping
                if len(str(cell.value)) > config['cell']['wrapThreshold']:
                    required_height = calculate_row_height(
                        cell.value, ws.column_dimensions[get_column_letter(col)].width,
                        config['cell']['wrapThreshold'], config['header']['fontSize']
                    )
                    ws.row_dimensions[row].height = required_height
                else:
                    ws.row_dimensions[row].height = None  

    # For non-merged header cells (just in case)
    for row in range(2, row_skip + 1):
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if not any(cell.coordinate in merged_cell_range for merged_cell_range in ws.merged_cells.ranges):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

                # Adjust height based on text wrapping
                if len(str(cell.value)) > config['cell']['wrapThreshold']:
                    required_height = calculate_row_height(
                        cell.value, ws.column_dimensions[get_column_letter(col)].width,
                        config['cell']['wrapThreshold'], config['header']['fontSize']
                    )
                    ws.row_dimensions[row].height = 50
                else:
                    ws.row_dimensions[row].height = None  # Auto-adjust if text fits within the threshold
    # Add body rows with styles
    for r, row in enumerate(main_df.values, start=row_skip + 1):
        for c, value in enumerate(row[1:], start=2):  # Skip the first column
            cell = ws.cell(row=r, column=c)
            # Preserve existing value or formula
            cell.alignment = alignment
            cell.border = border

            # Keep existing font bold for formula cells
            if cell.font.bold:  
                cell.font = Font(
                    color=body_font.color,
                    size=body_font.size,
                    name=body_font.name,
                    bold=True
                )
            else:
                cell.font = body_font

            # Check the length for wrapping and adjust alignment if needed
            if len(str(value)) > config['cell']['wrapThreshold']:
                cell.alignment = Alignment(
                    horizontal=config['alignment']['horizontal'],
                    vertical=config['alignment']['vertical'],
                    wrap_text=True  # Enable text wrapping
                )

            # Alternate row colors
            if r % 2 == 0:
                cell.fill = PatternFill(
                    start_color=config['body']['evenRowBackgroundColor'],
                    end_color=config['body']['evenRowBackgroundColor'],
                    fill_type='solid'
                )
            else:
                cell.fill = PatternFill(
                    start_color=config['body']['oddRowBackgroundColor'],
                    end_color=config['body']['oddRowBackgroundColor'],
                    fill_type='solid'
                )

    last_row = ws.max_row
    ws.delete_rows(last_row)

    wb.save(output_excel_file)
    print(f"Excel file saved as {output_excel_file}.")

    wb = openpyxl.load_workbook(xlsx_file_path)
    ws = wb.active  

    output = io.StringIO()
    output.write('<table cellpadding="0" cellspacing="0" style="border-collapse: collapse;">')

    # Keep track of merged cells that we have already processed
    processed_cells = set()

    for r in range(1, ws.max_row + 1):
        output.write(f'<tr style="height: auto;">')
        
        # Start from the second column to skip the first column
        for c in range(2, ws.max_column + 1):
            if (r, c) in processed_cells:
                continue  # Skip cells that are part of a merged range already processed

            cell = ws.cell(row=r, column=c)
            col_width = 'auto'  # Convert Excel units to approximate pixels

            # Default rowspan and colspan to 1
            rowspan = 1
            colspan = 1

            # Check if the current cell is in a merged cell range
            for merged_range in ws.merged_cells.ranges:
                cr = CellRange(str(merged_range))
                if cr.min_row <= r <= cr.max_row and cr.min_col <= c <= cr.max_col:
                    # Calculate rowspan and colspan based on the merged range
                    rowspan = cr.max_row - cr.min_row + 1
                    colspan = cr.max_col - cr.min_col + 1

                    # Mark all cells in the merged range as processed
                    for merged_r in range(cr.min_row, cr.max_row + 1):
                        for merged_c in range(cr.min_col, cr.max_col + 1):
                            processed_cells.add((merged_r, merged_c))
                    break

            # Build the cell style
            style = (
                f'style="padding: {config["cell"]["padding"]}px; '
                f'font-family: {config["font"]["family"]}; '
                f'width: {col_width}; '
            )

            # Add text wrapping style based on config
            if config['alignment'].get('wrapText', False):
                style += 'white-space: normal; '
            else:
                style += 'white-space: nowrap; overflow: hidden; text-overflow: ellipsis; '

            # Determine if the row is part of the header (controlled by row_skip)
            if r <= row_skip:
                style += (
                    f'background-color: #{config["header"]["backgroundColor"]}; '
                    f'color: #{config["header"]["fontColor"]}; '
                    f'font-weight: bold; '
                    f'font-size: {config["header"]["fontSize"]}pt; '
                    'text-align: center; '
                )
            else:  # Body rows style
                bg_color = (
                    config["body"]["evenRowBackgroundColor"]
                    if r % 2 == 0
                    else config["body"]["oddRowBackgroundColor"]
                )
                style += (
                    f'text-align: {config["alignment"]["horizontal"]}; 'f'text-align: {config["alignment"]["horizontal"]}; '
                    f'background-color: #{bg_color}; '
                    f'color: #{config["body"]["fontColor"]}; '
                    f'font-size: {config["body"]["fontSize"]}pt; '
                )

            # Add borders if enabled
            if config['cell'].get('borders', True):
                style += (
                    f'border: solid 1px grey; '
                )
            else:
                style += 'border: none; '

            # Set text alignment
            style += (
                f'vertical-align: {config["alignment"]["vertical"]}; '
                'word-wrap: break-word; '
                f'max-width: {config["cell"]["wrapThreshold"]*50}px; '
                '"'
            )
            
            if cell.data_type == "f":
                value=round(float(evaluate_formula(cell.value, ws)), 2)
                cell.value=value
            else: value = cell.value
            format=""
            if cell.number_format!='General':
                format=cell.number_format[4:]
            value=f"{value}{format}"

            cell_html = f'<td {style}'
            if colspan > 1:
                cell_html += f' colspan="{colspan}"'
            if rowspan > 1:
                cell_html += f' rowspan="{rowspan}"'
            cell_html += f'>{value}</td>'

            output.write(cell_html)
        output.write('</tr>')
    output.write('</table>')

    return output.getvalue()

def copy_table_to_clipboard(df, config_file='table_export_config.json', output_excel_file="output.xlsx"):
    config = load_config(config_file)
    html = excel_to_html_with_style(df, config)
    pyperclip.copy(html)
    print("Table copied to clipboard with styling. You can now paste it into Excel.")


