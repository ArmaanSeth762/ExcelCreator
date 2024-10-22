import pandas as pd
import algo
import re
from importlib import reload
reload(algo)
from algo import *
from io import StringIO
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import logging
logging.basicConfig(level=logging.INFO)

def remove_rows_above_main_data(df, threshold=0.3):
    """
    Remove rows above the main data based on a threshold of populated cells.
    """
    total_columns = len(df.columns)
    required_populated_cells = threshold * total_columns

    for idx, (index, row) in enumerate(df.iterrows()):
        populated_cells = row.count()
        if populated_cells >= required_populated_cells:
            # Main data starts here
            df = df.iloc[idx:].reset_index(drop=True)
            break
    return df

def remove_empty_left_columns(df, threshold=0.5):
    """
    Remove empty columns on the left until a column meets the populated cell threshold.
    """
    total_rows = len(df)
    required_populated_cells = threshold * total_rows

    columns = df.columns.tolist()
    for col in columns:
        populated_cells = df[col].count()
        if populated_cells >= required_populated_cells:
            # Main data column found
            break
        else:
            # Drop the column
            df = df.drop(columns=col)
    return df

def merge_insufficient_columns(df, threshold=0.7):
    """
    Merge columns with insufficient data into the next column.
    TODO: improve robustness of this function
    """
    total_rows = len(df)
    required_populated_cells = threshold * total_rows

    columns = df.columns.tolist()
    i = 0
    columns_to_drop = []

    while i < len(columns) - 1:
        col = columns[i]
        populated_cells = df[col].count()
        if populated_cells < required_populated_cells:
            next_col = columns[i + 1]
            # Combine current column into the next column
            df[next_col] = df[col].fillna('') + ' ' + df[next_col].fillna('')
            columns_to_drop.append(col)
            i += 1  # Move to next column
        else:
            break
    df = df.drop(columns=columns_to_drop)
    return df

def remove_empty_rows_and_columns(df):
    """
    Remove rows and columns that are entirely empty.
    """
    df = df.dropna(axis=0, how='all')  # Drop empty rows
    df = df.dropna(axis=1, how='all')  # Drop empty columns
    return df

def general_data_cleaning(df):
    """
    Perform general data cleaning operations.
    """
    # Apply strip only to string columns
    for col in df.columns:
        # Check if the column contains any string values
        if df[col].apply(lambda x: isinstance(x, str)).any():
            df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
    return df

def clean_dataframe(df):
    if df.empty:
        raise ValueError("The DataFrame is empty.")
    try:
        df = remove_rows_above_main_data(df)
        # logging.info("Removed rows above main data.")
    except Exception as e:
        logging.error(f"Error removing rows: {e}")

    try:
        df = remove_empty_left_columns(df)
        # logging.info("Removed empty left columns.")
    except Exception as e:
        logging.error(f"Error removing columns: {e}")

    try:
        df = merge_insufficient_columns(df)
        # logging.info("Merged insufficient columns.")
    except Exception as e:
        logging.error(f"Error merging columns: {e}")

    try:
        df = remove_empty_rows_and_columns(df)
        # logging.info("Removed empty rows and columns.")
    except Exception as e:
        logging.error(f"Error removing empty rows/columns: {e}")

    try:
        df = general_data_cleaning(df)
        # logging.info("Performed general data cleaning.")
    except Exception as e:
        logging.error(f"Error in general data cleaning: {e}")
    df.reset_index(drop=True, inplace=True)
    return df

def convert_to_float(data):
    if isinstance(data, pd.Series):
        data=data.to_string()

    if data is None or pd.isna(data):
        return None, None
    elif isinstance(data, (float, int)):
        return float(data), None
    elif isinstance(data, str):
        data = ' '.join(data.strip().split()).replace(" ","")
        value=data
        currency_symbols = ['$', '€', '£', '¥', '₹', '₽', '₩', '₪', 'Fr', 'kr', 'R$', 'A$', 'C$', '₺', '₫', '₦']
        symbol = None
        
        if "%"  in value:
            symbol = "%"
            value = value.replace("%", "")
        elif "percent"  in value:
            symbol = "%"
            value = value.replace("%", "")
        else:
            for curr_symbol in currency_symbols:
                if curr_symbol in value:
                    symbol = curr_symbol
                    value = value.replace(curr_symbol, "")
                    break
        
        value = value.replace(",", "").replace("'", "")
        if value == "-":
            return None, None
        if "(" in value and ")" in value:
            value = "-" + value.strip("()")

        try:
            value=float(value)
            if symbol=="%":
                if abs(value)>1:
                    value=value/100
            return value, symbol

        except ValueError:
            return data, None  # Return the original string if it can't be converted
    else:
        return None, None

def preprocess_dataframe(df):
    """
    Preprocess the DataFrame for formula calculations.
    Args:
    df (pd.DataFrame): Input DataFrame
    Returns:
    tuple: (list of bool indicating text-based columns, matrix of symbols, preprocessed DataFrame)
    """

    # Create a matrix of symbols and preprocess the DataFrame
    symbol_matrix =[]
    threshold=0.3

    # Backup original columns
    original_df = df.copy()
    is_text_based = [False] * len(df.columns)  # Initialize all columns as not text-based
    for i, col in enumerate(df.columns):
        symbol_col = []
        
        for index in range(len(df)):
            value, symbol = convert_to_float(df.at[index, col])
            symbol_col.append(symbol)
            df.at[index, col] = value
        
        # After the end of the row loop, determine if the column is text-based
        if sum(isinstance(x, str) or x is None for x in df[col].values) > threshold * len(symbol_col):
            for index in range(len(df)):
                if df.at[index, col] is None or isinstance(df.at[index, col], str):
                    if symbol_col[index]!='date':
                        if isinstance(original_df.at[index, col], pd.Series):
                            df.at[index, col] = original_df.at[index, col].to_string()
                        else:
                            df.at[index, col] = original_df.at[index, col]  # Replace with original value if None or string
                        symbol_col[index]=None
            is_text_based[i] = True  # Column remains text-based if any value is None or string
        else:
            is_text_based[i] = False  # Mark column as not text-based
        symbol_matrix.append(symbol_col)
    symbol_matrix = [list(row) for row in zip(*symbol_matrix)]
    symbol_matrix=[[None]*len(df.columns)]+symbol_matrix
    return is_text_based, symbol_matrix, df

def checkval(str1, str2) -> bool:
    if not isinstance(str1, str) or not isinstance(str2, str):
        return False
    # Regular expression to match strings with optional numeric suffixes after a dot
    pattern = r'^(.+?)(\.\d+)?$'
    
    # Match the pattern for both strings
    match1 = re.match(pattern, str1)
    match2 = re.match(pattern, str2)
    
    if match1 and match2:
        # Compare the parts before the dot (ignore the numeric suffix if present)
        return match1.group(1) == match2.group(1)
    
    # If the pattern is not matched, perform exact equality check
    return str1 == str2

def save_excel(file_path, output_path, row_skip, formula_sign, formula_till, is_text_based, symbol_matrix):
    # Load the workbook and get the active sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    max_row = sheet.max_row
    max_col = sheet.max_column

    # Define a bold font style
    bold_font = Font(bold=True)
    
    # Define wrap text alignment for both rows and columns
    wrap_alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    # Iterate over all cells in the sheet
    for row_sheet in range(1, max_row + 1):
        for col_sheet in range(2, max_col + 1):
            cell = sheet.cell(row=row_sheet, column=col_sheet)
            cell.alignment = wrap_alignment

    # Track rows that have formulas
    for r in range(1, row_skip+1):
        # print("r=",r)
        # First, merge cells row-wise (left to right)
        start_cell = 2
        start_cell_value = sheet.cell(row=r, column=start_cell).value
        col = 3
        if sheet.cell(row=r, column=2).coordinate in sheet.merged_cells:
            merged_range = sheet.merged_cells.ranges
            for merged in merged_range:
                if merged.bounds[0] == start_cell and merged.bounds[1] <= r <= merged.bounds[3]:
                    col = merged.bounds[2] + 1
                    break  # Exit the loop after finding the merged range
        # if r==2: print(start_cell, start_cell_value)
        while col <= max_col:
            current_cell = sheet.cell(row=r, column=col)
            # if r==2: print(col, current_cell.value)
            if not checkval(current_cell.value, start_cell_value):
                
                # Merge cells if the current cell value is the same as the left cell
                if start_cell < col - 1 and not sheet.cell(row=r, column=start_cell).coordinate in sheet.merged_cells:  # Ensure we only merge if there's a range to merge and it's not already merged
                    # Check if the merge is correct by comparing the values of the cells to be merged
                    merge_correct = True
                    for merge_col in range(start_cell, col):
                        if sheet.cell(row=r, column=merge_col).value != start_cell_value:
                            merge_correct = False
                            break
                    if merge_correct:
                        # if r==2: print("Merge:",start_cell, "-",col-1)
                        sheet.merge_cells(start_row=r, start_column=start_cell, end_row=r, end_column=col - 1)
                        # print(sheet.merged_cells.ranges)
                    # else:
                        # if r==2: print("Merge not performed: values differ.")
                
                start_cell = col
                start_cell_value = current_cell.value
                # if r == 2: print("Start Cell:", start_cell, "Value:", start_cell_value, type(start_cell_value))
                if current_cell.coordinate in sheet.merged_cells:
                    merged_range = sheet.merged_cells.ranges
                    for merged in merged_range:
                        if merged.bounds[0] == col and merged.bounds[1] <= r <= merged.bounds[3]:
                            col = merged.bounds[2]  # Set col to the last column of the merged group
                            break
            else:
                current_cell.value = start_cell_value
                # if r==2: print("Check", current_cell.value, start_cell_value)
            col += 1
        
        # Then, merge with the previous row if necessary
        if r > 1:
            col=2
            while(col<=max_col):
                current_cell = sheet.cell(row=r, column=col)
                above_cell = sheet.cell(row=r - 1, column=col)
                above_cell_value=above_cell.value
                flag=False
                if above_cell.coordinate in sheet.merged_cells:
                    # Get the top-left cell of the merged range
                    merged_ranges = sheet.merged_cells.ranges
                    for merged in merged_ranges:
                        if merged.bounds[0] <= col <= merged.bounds[2] and merged.bounds[1] <= r-1:
                            top_left_cell = sheet.cell(row=merged.bounds[1], column=merged.bounds[0])
                            above_cell_value = top_left_cell.value
                            flag=True
                            break
                if current_cell.value == above_cell_value:
                    # Check if the above cell is part of a merged group
                    if flag:
                        merged_ranges = sheet.merged_cells.ranges
                        for merged in merged_ranges:
                            if merged.bounds[0] <= col <= merged.bounds[2] and merged.bounds[1] <= r- 1:
                                print(str(merged))
                                print(merged.bounds)
                                sheet.unmerge_cells(str(merged))
                                
                                lower_right_col=col
                                lower_right_row=r
                                if current_cell.coordinate in sheet.merged_cells:
                                    merged_ranges = sheet.merged_cells.ranges
                                    for merged_curr in merged_ranges:
                                        if merged_curr.bounds[0] <= col <= merged_curr.bounds[2] and merged_curr.bounds[1] <= r <= merged_curr.bounds[3]:
                                            # Find the lower right cell of the merged_curr range
                                            
                                            lower_right_row = merged_curr.bounds[3]
                                            lower_right_col = merged_curr.bounds[2]
                                            sheet.unmerge_cells(str(merged_curr))
                                            break
                                # Merge the new diagonal cells
                                sheet.merge_cells(start_row=merged.bounds[1], start_column=merged.bounds[0], 
                                                  end_row=lower_right_row, end_column=lower_right_col)
                                col=lower_right_col
                                break
                    else:
                        # Simple merge if above cell is not part of a merged group
                        start_row = r - 1
                        sheet.merge_cells(start_row=start_row, start_column=col, end_row=r, end_column=col)
                col+=1

    # Merge cells for text-based columns
    for col_sheet in range(2, max_col + 1):
        col_index = col_sheet - 2  # Adjusted to account for all columns
        # Check if the column is text-based
        if is_text_based[col_index]:
            start_row = row_skip + 1
            start_row_val = sheet.cell(row=start_row, column=col_sheet).value
            curr = start_row + 1
            while curr <= max_row:
                curr_cell = sheet.cell(row=curr, column=col_sheet)
                if curr_cell.value != start_row_val:
                    if curr - 1 > start_row:
                        sheet.merge_cells(start_row=start_row, start_column=col_sheet, end_row=curr - 1, end_column=col_sheet)
                    start_row = curr
                    start_row_val = curr_cell.value
                curr += 1
            
            if start_row < curr - 1:
                sheet.merge_cells(start_row=start_row, start_column=col_sheet, end_row=curr - 1, end_column=col_sheet)
            
    # Track rows that have formulas
    rows_with_formulas = []
    for col_sheet in range(2, max_col + 1):
        col = get_column_letter(col_sheet)
        col_index = col_sheet - 2
        if not is_text_based[col_index]:
            first_cell_empty = sheet[col + str(row_skip + 1)].value is None or sheet[col + str(row_skip + 1)].value == ""
            for row_sheet in range(row_skip + 1, max_row + 1):
                row = row_sheet - (row_skip + 1)
                cell = sheet[col + str(row_sheet)]

                if formula_till>0 and len(formula_sign[row]) > 0:
                    generated_formula = "="  # Initialize the formula string
                    for i in formula_sign[row]:
                        # Check if the referred cell value is 0 and the first cell is empty
                        referred_cell = sheet[col + str(abs(i) + row_skip)]
                        if referred_cell.value == 0 and first_cell_empty:
                            continue  # Skip this cell reference in the formula
                        
                        if generated_formula != "=":
                            generated_formula += " + " if i > 0 else " - "
                        elif i < 0:
                            generated_formula += "-"
                        generated_formula += col
                        generated_formula += str(abs(i) + row_skip)
                    if generated_formula != "=":
                        cell.value = generated_formula  # Set the cell value to the generated formula
                        rows_with_formulas.append(row_sheet)  # Track the row with a formula
                    else:
                        cell.value = 0  # Set to 0 if all references were removed
                else:
                    cell.value = convert_to_float(cell.value)[0]

                if symbol_matrix[row+2][col_index+1]:
                    cell.number_format = f'0.00{symbol_matrix[row+2][col_index+1]}'
            formula_till-=1
    
        sheet.column_dimensions[col].width = 15
        
    for row in rows_with_formulas:
        for col_row in range(2, max_col + 1):
            sheet.cell(row=row, column=col_row).font = bold_font

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    cell.value = cell.value

    print("File successfully saved!")
    workbook.save(output_path)  # Save the workbook to the specified output path
    workbook.close()  # Close the workbook

def html_to_dataframe(html):
    df_list = pd.read_html(StringIO(html))  
    if df_list:
        df = df_list[0]
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = pd.MultiIndex.from_tuples([tuple([col if 'Unnamed' not in col else "-" for col in cols]) for cols in df.columns.values.tolist()])
        else:
            if df.columns.values.tolist()==[i for i in range (0, len(df.columns))]:
                df.columns = df.iloc[0]
                df.columns = [col if not isinstance(col, str) or 'Unnamed' not in col else "-" for col in df.columns.values.tolist()]
                df = df[1:]
        return df
    else:
        raise ValueError("No tables found in the provided HTML string.")
    
def save_double_column_df(df, xl_writer, startrow = 0, **kwargs):
    '''Function to save doublecolumn DataFrame, to xlwriter'''
    # inputs:
    # df - pandas dataframe to save
    # xl_writer - book for saving
    # startrow - row from wich data frame will begins
    # **kwargs - arguments of `to_excel` function of DataFrame`
    df.drop(df.index).to_excel(xl_writer, startrow = startrow, **kwargs)
    df.to_excel(xl_writer, startrow = startrow + 1, header = False, **kwargs)

def run(html):
    df = html_to_dataframe(html)
    row_skip = max([len(col) for col in df.columns]) if isinstance(df.columns, pd.MultiIndex) else 1
    df=clean_dataframe(df)
    is_text_based, symbol_matrix, df = preprocess_dataframe(df)
    symbol_matrix = [[None] * (len(symbol_matrix[0]) + 1)] + [[None] + row for row in symbol_matrix]
    col_skip=0
    while(col_skip<len(is_text_based) and is_text_based[col_skip]): col_skip+=1
    all_columns = []
    for column in df.columns:
        ls = df[column].values.tolist()
        for i in range(len(ls)):
            if isinstance(ls[i], str):
                try:
                    ls[i] = float(ls[i].replace(',', '').replace(" ","").strip())  # Remove commas and convert to float
                except ValueError:
                    ls[i] = 0.0  # If conversion fails, set to 0.0
            elif pd.isna(ls[i]):
                ls[i] = 0.0
        all_columns.append(ls)
    all_columns=all_columns[col_skip:]
    formulas, formula_till =get_all_formulas(all_columns)
    
    if isinstance(df.columns, pd.MultiIndex):
        for formula in formulas:
            for f in formula:
                tmp=abs(f)-1
                if f>0:
                    f=tmp
                else:
                    f=-tmp
    xl_writer = pd.ExcelWriter("tmp_.xlsx",engine='xlsxwriter')
    save_double_column_df(df, xl_writer)
    xl_writer.close()
    print(formula_till)
    save_excel('tmp_.xlsx', 'tmp.xlsx', row_skip, formulas, formula_till, is_text_based, symbol_matrix)
    return df, row_skip, col_skip, is_text_based, symbol_matrix, all_columns, formulas
